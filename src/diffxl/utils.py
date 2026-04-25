
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from typing import Optional
import datetime


def save_diff_report(
    df_added: pd.DataFrame,
    df_removed: pd.DataFrame,
    df_changed: pd.DataFrame,
    df_new: pd.DataFrame,
    key_col: str,
    output_path: str,
    df_old_dups: Optional[pd.DataFrame] = None,
    df_new_dups: Optional[pd.DataFrame] = None,
    old_file_name: str = "",
    new_file_name: str = "",
    df_old: Optional[pd.DataFrame] = None,
) -> None:
    """Save the difference report to an Excel file with formatted highlighting.

    Sheets:
        1. Summary       – Overview stats and per-column change counts.
        2. Added Rows     – New rows (green tab).
        3. Removed Rows   – Deleted rows (red tab).
        4. Changed Details – Cell-level changes (yellow tab).
        5. Full Diff - excl removals      – New file snapshot with highlighted changes and additions.
        6. Ignored Duplicates – (optional) if dedup was used.
    """

    # ── Styles ────────────────────────────────────────────────────────────────
    fill_green  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_red    = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    fill_stat_header = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")

    font_bold       = Font(bold=True)
    font_header     = Font(bold=True, color="FFFFFF", size=11)
    font_title      = Font(bold=True, size=14)
    font_subtitle   = Font(bold=True, size=11)
    font_stat_label = Font(bold=True, size=10)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left   = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Helper to write sheet
        def write_sheet(df: pd.DataFrame, name: str) -> None:
            """Write a DataFrame to a sheet, handling empty frames."""
            if df.empty:
                pd.DataFrame(columns=df.columns).to_excel(writer, sheet_name=name, index=False)
            else:
                df.to_excel(writer, sheet_name=name, index=False)

        # ── 1. Summary Sheet ─────────────────────────────────────────────────
        # Create a blank sheet first, we'll populate it manually.
        pd.DataFrame().to_excel(writer, sheet_name="Summary", index=False)

        # ── 2. Standard Diff Sheets ───────────────────────────────────────────
        write_sheet(df_added, "Added Rows")
        write_sheet(df_removed, "Removed Rows")
        write_sheet(df_changed, "Changed Details")

        # ── 3. Full Diff - excl removals Sheet ────────────────────────────────────────────────
        if df_new.empty:
            pd.DataFrame(columns=["Info"]).to_excel(writer, sheet_name="Full Diff - excl removals", index=False)
        else:
            # Start data at row 2 (leave row 1 for per-column change counts)
            df_new.to_excel(writer, sheet_name="Full Diff - excl removals", index=False, startrow=1)

        # ── 4. Ignored Duplicates Sheet (if any) ─────────────────────────────
        has_dups = (df_old_dups is not None and not df_old_dups.empty) or \
                   (df_new_dups is not None and not df_new_dups.empty)
        if has_dups:
            dups_report = []
            if df_old_dups is not None and not df_old_dups.empty:
                df_old_dups = df_old_dups.copy()
                df_old_dups["Duplicate Source"] = "Original File"
                dups_report.append(df_old_dups)

            if df_new_dups is not None and not df_new_dups.empty:
                df_new_dups = df_new_dups.copy()
                df_new_dups["Duplicate Source"] = "New File"
                dups_report.append(df_new_dups)

            if dups_report:
                df_combined_dups = pd.concat(dups_report, ignore_index=True)
                cols = df_combined_dups.columns.tolist()
                cols = ['Duplicate Source'] + [c for c in cols if c != 'Duplicate Source']
                df_combined_dups = df_combined_dups[cols]
                write_sheet(df_combined_dups, "Ignored Duplicates")

        # ── Access Workbook ───────────────────────────────────────────────────
        workbook = writer.book

        # ── Tab Colors ────────────────────────────────────────────────────────
        try:
            workbook["Summary"].sheet_properties.tabColor = "4472C4"
            workbook["Added Rows"].sheet_properties.tabColor = "00B050"
            workbook["Removed Rows"].sheet_properties.tabColor = "FF0000"
            workbook["Changed Details"].sheet_properties.tabColor = "FFC000"
            workbook["Full Diff - excl removals"].sheet_properties.tabColor = "0070C0"
            if has_dups:
                workbook["Ignored Duplicates"].sheet_properties.tabColor = "999999"
        except Exception:
            pass

        # ── Populate Summary Sheet ────────────────────────────────────────────
        ws_summary = workbook["Summary"]

        # Title
        ws_summary["A1"] = "DiffXL Report"
        ws_summary["A1"].font = font_title
        ws_summary["A2"] = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_summary["A2"].font = Font(italic=True, color="808080")

        # File info
        row = 4
        ws_summary.cell(row=row, column=1, value="Original File:").font = font_stat_label
        ws_summary.cell(row=row, column=2, value=old_file_name)
        row += 1
        ws_summary.cell(row=row, column=1, value="New File:").font = font_stat_label
        ws_summary.cell(row=row, column=2, value=new_file_name)
        row += 1
        ws_summary.cell(row=row, column=1, value="Key Column:").font = font_stat_label
        ws_summary.cell(row=row, column=2, value=key_col)

        # Overall stats
        row += 2
        ws_summary.cell(row=row, column=1, value="Summary Statistics").font = font_subtitle

        count_changed_rows = df_changed[key_col].nunique() if not df_changed.empty else 0
        total_old = len(df_old) if df_old is not None else 0
        total_new = len(df_new)

        stats = [
            ("Rows in Original File", total_old),
            ("Rows in New File", total_new),
            ("Added Rows", len(df_added)),
            ("Removed Rows", len(df_removed)),
            ("Changed Rows", count_changed_rows),
            ("Changed Cells", len(df_changed)),
            ("Unchanged Rows", total_new - len(df_added) - count_changed_rows),
        ]

        row += 1
        for label, value in stats:
            c1 = ws_summary.cell(row=row, column=1, value=label)
            c2 = ws_summary.cell(row=row, column=2, value=value)
            c1.font = font_stat_label
            c1.fill = fill_stat_header
            c2.alignment = align_center
            c2.border = thin_border
            c1.border = thin_border
            row += 1

        # Schema changes
        if df_old is not None:
            old_cols_set = set(df_old.columns)
            new_cols_set = set(df_new.columns)
            added_cols = sorted(new_cols_set - old_cols_set)
            removed_cols = sorted(old_cols_set - new_cols_set)

            if added_cols or removed_cols:
                row += 1
                ws_summary.cell(row=row, column=1, value="Column Changes").font = font_subtitle
                row += 1
                if added_cols:
                    ws_summary.cell(row=row, column=1, value="Added Columns:").font = font_stat_label
                    ws_summary.cell(row=row, column=2, value=", ".join(added_cols))
                    row += 1
                if removed_cols:
                    ws_summary.cell(row=row, column=1, value="Removed Columns:").font = font_stat_label
                    ws_summary.cell(row=row, column=2, value=", ".join(removed_cols))
                    row += 1

        # Per-column change counts
        row += 1
        ws_summary.cell(row=row, column=1, value="Changes per Column").font = font_subtitle
        row += 1

        # Header row
        for ci, header in enumerate(["Column Name", "Changed Cells"], start=1):
            c = ws_summary.cell(row=row, column=ci, value=header)
            c.font = font_header
            c.fill = fill_header
            c.alignment = align_center
            c.border = thin_border
        row += 1

        if not df_changed.empty:
            col_counts = df_changed["Column"].value_counts().sort_values(ascending=False)
            for col_name, count in col_counts.items():
                c1 = ws_summary.cell(row=row, column=1, value=col_name)
                c2 = ws_summary.cell(row=row, column=2, value=count)
                c1.border = thin_border
                c2.border = thin_border
                c2.alignment = align_center
                row += 1
        else:
            ws_summary.cell(row=row, column=1, value="No cell changes detected.")
            row += 1

        # Auto-fit Summary column widths
        ws_summary.column_dimensions["A"].width = 28
        ws_summary.column_dimensions["B"].width = 40

        # ── Style data sheet headers ──────────────────────────────────────────
        def style_data_sheet(ws, df: pd.DataFrame, header_row: int = 1) -> None:
            """Apply header styling, freeze panes, and auto-fit widths."""
            if df.empty or ws.max_column == 0:
                return

            # Style header row
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=header_row, column=col_idx)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border

            # Freeze below header
            ws.freeze_panes = f"A{header_row + 1}"

            # Auto-filter dropdowns on header row
            last_col = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A{header_row}:{last_col}{ws.max_row}"

            # Auto-fit column widths (sampling first 100 data rows + header)
            data_start = header_row + 1
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = len(str(ws.cell(row=header_row, column=col_idx).value or ""))
                for row_idx in range(data_start, min(ws.max_row + 1, data_start + 100)):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val is not None:
                        max_len = max(max_len, len(str(val)))
                # Clamp between 8 and 40
                ws.column_dimensions[col_letter].width = min(40, max(8, max_len + 2))

        style_data_sheet(workbook["Added Rows"], df_added)
        style_data_sheet(workbook["Removed Rows"], df_removed)
        style_data_sheet(workbook["Changed Details"], df_changed)

        if not df_new.empty:
            style_data_sheet(workbook["Full Diff - excl removals"], df_new, header_row=2)

        if has_dups and "Ignored Duplicates" in workbook.sheetnames:
            style_data_sheet(workbook["Ignored Duplicates"], df_combined_dups)

        # ── Highlight Full Diff - excl removals sheet ─────────────────────────────────────────
        ws = workbook["Full Diff - excl removals"]

        # Per-column change counts in row 1
        col_change_counts = {}
        if not df_changed.empty:
            col_change_counts = df_changed["Column"].value_counts().to_dict()

        fill_count_bg = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        font_count = Font(bold=True, size=9, color="666666")

        for ci, col_name in enumerate(df_new.columns):
            excel_col = ci + 1
            count = col_change_counts.get(col_name, 0)
            cell = ws.cell(row=1, column=excel_col)
            if count > 0:
                cell.value = f"{count} changes"
            else:
                cell.value = ""
            cell.font = font_count
            cell.fill = fill_count_bg
            cell.alignment = align_center

        # Row offset: row 1 = counts, row 2 = header, row 3+ = data
        DATA_ROW_OFFSET = 3

        key_map = {}
        for idx, val in enumerate(df_new[key_col]):
            k = str(val).strip()
            key_map[k] = idx

        col_map = {str(c): i for i, c in enumerate(df_new.columns)}

        # Highlight Added Rows (Green)
        if not df_added.empty:
            for val in df_added[key_col]:
                k = str(val).strip()
                if k in key_map:
                    row_idx = key_map[k]
                    excel_row = row_idx + DATA_ROW_OFFSET
                    max_col = len(df_new.columns)
                    for c_idx in range(1, max_col + 1):
                        cell = ws.cell(row=excel_row, column=c_idx)
                        cell.fill = fill_green
                        cell.comment = Comment(f"Added row", "Diffxl")

        # Highlight Changed Cells (Yellow) + mark UID column
        if not df_changed.empty:
            changed_keys = set()
            for _, row in df_changed.iterrows():
                k = str(row[key_col]).strip()
                col_name = str(row["Column"])

                if k in key_map and col_name in col_map:
                    changed_keys.add(k)
                    row_idx = key_map[k]
                    col_idx = col_map[col_name]
                    excel_row = row_idx + DATA_ROW_OFFSET
                    excel_col = col_idx + 1
                    cell = ws.cell(row=excel_row, column=excel_col)
                    cell.fill = fill_yellow
                    old_val = str(row["Old Value"])
                    cell.comment = Comment(f"{old_val}", "Old value")

            # Mark UID column for changed rows (red font — enables Filter by Color)
            key_excel_col = col_map.get(key_col, 0) + 1
            for k in changed_keys:
                if k in key_map:
                    excel_row = key_map[k] + DATA_ROW_OFFSET
                    cell = ws.cell(row=excel_row, column=key_excel_col)
                    cell.font = cell.font.copy(color="FF0000")
                    cell.comment = Comment("Changes made to this row", "Diffxl")
